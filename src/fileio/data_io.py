import os
from os.path import isfile, join, isdir
import numpy as np
import openpyxl
from .data_processing import train_test_transpose
from .data_processing import data_collection

################################################################
#File IO

def check_file_exists(data_file):
    return isfile(data_file)

def listall_withinsubdir(dir):
    return os.walk(dir)

def list_files(dir):
    onlyfiles = [f for f in os.listdir(dir) if isfile(join(dir, f))]
    return sorted(onlyfiles)

def listall(dir):
    return os.listdir(dir)

def list_sub_folders(dir):
    sub_folders = [f for f in os.listdir(dir) if isdir(join(dir,f))]
    return sorted(sub_folders)


def delete_files_startwith(dir, start_str):
    all_files = list_files(dir)
    for file in all_files:
        if file.startswith(start_str):
            os.remove(os.path.join(dir, file))

def check_folder_exists(data_folder):
    try:
        os.stat(data_folder)
    except:
        print (data_folder + ' does not exists!')
        return False
    return True

def check_file_exists(data_file):
    return isfile(data_file)


# Check folder exists or not, create folders if not exists
def init_folder(data_folder):
    split_key = '/'
    if data_folder.endswith(split_key):
        data_folder = data_folder[:-1]
    folder_array = data_folder.split(split_key)
    data_folder = ''
    for item in folder_array:
        data_folder = data_folder + item + split_key
        if item == '..':
            continue
        try:
            os.stat(data_folder)
        except:
            os.mkdir(data_folder)

    return data_folder




#End of file IO
################################################################

def file_read_split(file_name, class_column=0, delimiter=' ', header=True):
    data_matrix, attr_num = file_reading(file_name, delimiter, header)
    x_matrix, y_vector = x_y_spliting(data_matrix, class_column)
    x_row, x_col = x_matrix.shape
    attr_len = x_col/attr_num
    x_matrix = x_matrix.reshape(x_row, attr_num, attr_len)
    return x_matrix, y_vector


##
# Read the giving file and store the data into matrix structure
# the return format is folat, and the minimal value of y may not be 0
def file_reading(file_name, delimiter=' ', header=True):
    num = 0
    data_matrix = []
    header_line = "-1"
    with open(file_name) as f:
        data_row = []
        for line in f:
            if header is True:
                header = False
                header_line = line.strip()
                continue
            num = num + 1
            #if num > 100:
            #    break
            data_row = line.split(delimiter)
            data_matrix.append(data_row)
    row_num = len(data_matrix)
    col_num = len(data_matrix[0])
    data_matrix = np.array(data_matrix, dtype=float) #.reshape(row_num, col_num)
    data_matrix.astype(float)
    header_line = int(header_line)
    return data_matrix, header_line

##
# Read the giving file and store the data into matrix structure
# the return format is folat, and the minimal value of y may not be 0
def file_writing(data_matrix, file_name, attr_num=-1, delimiter=' '):
    data_row, data_col = data_matrix.shape
    with open(file_name, 'w') as f:
        if attr_num > 0:
            f.write(str(int(attr_num)) + '\n')
        for row in range(0, data_row):
            row_vector = data_matrix[row, :]
            row_str = str(int(row_vector[0]))
            for index in range(1, data_col):
                row_str = row_str + delimiter + str(row_vector[index])
            f.write(row_str + '\n')
    
def file_writingxy(data_x_matrix, data_y_vector, file_name, attr_num=-1, delimiter=' '):
    data_row, data_col = data_x_matrix.shape
    with open(file_name, 'w') as f:
        if attr_num > 0:
            f.write(str(int(attr_num)) + '\n')
        for row in range(0, data_row):
            row_vector = data_x_matrix[row, :]
            row_label = str(int(data_y_vector[row]))
            row_str = row_label
            for index in range(0, data_col):
                row_str = row_str + delimiter + str(row_vector[index])
            f.write(row_str + '\n')

##
# from a data_matrix, split the x_matrix and y_vector based on class_column
# values in y_vector can not be negative
# returned y_vector with minimal value is 0
def x_y_spliting(data_matrix, class_column):
    y_vector = data_matrix[:, class_column].astype(int)
    y_vector = y_vector - min(y_vector)
    x_matrix = np.delete(data_matrix, class_column, 1)
    return x_matrix, y_vector



# def train_test_file_reading(train_file, test_file, class_column=0, delimiter=' ', header=True):
#     train_matrix, attr_num = file_reading(train_file, delimiter, header)
#     test_matrix, attr_num = file_reading(test_file, delimiter, header)
#     train_x_matrix, train_y_vector = x_y_spliting(train_matrix, class_column)
#     test_x_matrix, test_y_vector = x_y_spliting(test_matrix, class_column)
#     train_min_class = min(train_y_vector)
#     test_min_class = min(test_y_vector)
#     if train_min_class != 0 or test_min_class !=0:
#         raise Exception("minimum class does not match")
#     return train_x_matrix, train_y_vector, test_x_matrix, test_y_vector

def train_test_file_reading(train_file, test_file='', valid_file='', class_column=0, delimiter=' ', header=True):
    train_matrix, attr_num = file_reading(train_file, delimiter, header)
    train_x_matrix, train_y_vector = x_y_spliting(train_matrix, class_column)
    data_group = data_collection(train_x_matrix, train_y_vector, class_column)
    if test_file == '':
        test_x_matrix = None
        test_y_vector = None
    else:
        test_matrix, attr_num = file_reading(test_file, delimiter, header)
        test_x_matrix, test_y_vector = x_y_spliting(test_matrix, class_column)
        data_group.test_x_matrix = test_x_matrix
        data_group.test_y_vector = test_y_vector
    if valid_file == '':
        valid_x_matrix = None
        valid_y_vector = None
    else:
        valid_matrix, attr_num = file_reading(valid_file, delimiter, header)
        valid_x_matrix, valid_y_vector = x_y_spliting(valid_matrix, class_column)
        data_group.valid_x_matrix = valid_x_matrix
        data_group.valid_y_vector = valid_y_vector
    return data_group, int(attr_num)


# convert 2d x matrix to 4d matrix
# trans is True: ins * attr_len * attr_num * 1
# trans is False: ins * attr_num * attr_len * 1
def data_group_processing(data_group, attr_num, trans=True):
    train_x_matrix = data_group.train_x_matrix
    attr_len = -1
    if train_x_matrix is not None:
        train_row, train_col = train_x_matrix.shape
        attr_len = int(train_col/attr_num)
        data_group.train_x_matrix = train_test_transpose(train_x_matrix, attr_num, attr_len, trans)

    test_x_matrix = data_group.test_x_matrix
    if test_x_matrix is not None:
        test_row, test_col = test_x_matrix.shape
        if attr_len > 0:
            if attr_len != test_col/attr_num:
                raise Exception("data set attr len not match!!")
        else:
            attr_len = test_col/attr_num
        data_group.test_x_matrix = train_test_transpose(test_x_matrix, attr_num, attr_len, trans)

    valid_x_matrix = data_group.valid_x_matrix
    if valid_x_matrix is not None:
        valid_row, valid_col = valid_x_matrix.shape
        if attr_len > 0:
            if attr_len != valid_col/attr_num:
                raise Exception("data set attr len not match!!")
        else:
            attr_len = valid_col/attr_num
        data_group.valid_x_matrix = train_test_transpose(valid_x_matrix, attr_num, attr_len, trans)
        


def write_to_excel(data_matrix, excel_file, start_row=1, start_col=1, sheet_name='sheet1'):
    try:
        workbook = openpyxl.load_workbook(excel_file)
        worksheet = workbook.get_sheet_by_name(sheet_name)
    except:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = sheet_name
        #worksheet = workbook.create_sheet(sheet_name)
    
    for row, row_value in enumerate(data_matrix):
        for col, col_value in enumerate(row_value):
            try:
                col_value = float(col_value)
            except:
                pass
            worksheet.cell(row=row+start_row, column=col+start_col).value = col_value
    workbook.save(excel_file)



def data_dict_loading(data_folder, data_keyword, class_column=0, delimiter=' ', header=True):
    data_file_list = list_files(data_folder)
    data_dict = {}
    for data_file in data_file_list:
        if data_keyword not in data_file:
            continue
        fold_num = int(data_file.split('_')[1])
        train_matrix, attr_num = file_reading(data_folder + data_file, delimiter, header)
        train_x_matrix, train_y_vector = x_y_spliting(train_matrix, class_column)
        data_dict[fold_num] = [train_x_matrix, train_y_vector]
    return data_dict

def softmax(x_vector):
    theta = 2.0
    ps = np.exp(x_vector * theta)
    ps /= np.sum(ps)
    return ps


if __name__ == '__main__':
    #data_folder = '../../data/arc_activity_recognition/s1_ijcal_10folds/'
    #file_name_1 = "train.txt"
    #file_name_2 = "test.txt"
    #out_file = "all.txt"
    #merge_two_files(data_folder, file_name_1, file_name_2, out_file)

    # obj_folder = "../../object/dsa/train_test_3_fold/"
    # print (list_sub_folders(obj_folder))
    # sdf
    # attr_num = 2
    # attr_len = 3
    # a = np.random.rand(2, attr_num*attr_len)
    # b = train_test_transpose(a, attr_num, attr_len)
    # print ("a")
    # print (a.reshape(2, attr_num, attr_len))
    # print ("b")
    # print (b.reshape(2, attr_len, attr_num))

    a = [2.7139020e-01, 1.7380084e+00, 3.2913544e+00, 6.3995056e+00, 3.5484326e+00, 1.4874511e+00, 4.4787264e-01, 1.4647619e-01, 8.8376619e-02, 6.2849522e-02, 1.4465781e-02, 3.0703414e-02, 4.8740126e-02, 5.0060004e-02, 6.2272283e-03, 1.2495031e-03, 9.6714654e-04, 1.5672009e-03, 1.6876594e-03, 1.8615086e-03]
    a= [9.570305 , 42.39435  , 16.143522 ,  3.5946317,  6.6580033,
        6.338355 ,  2.963811 ,  1.6737944,  1.0775504,  2.6309962,
        1.325895 ,  1.8399512,  3.1222334,  5.617679 , 25.377987 ,
       11.16932  ,  4.3038325,  5.3518367,  6.9019294,  2.6773973]
    a = [ 8.889062  ,  5.035801  ,  3.5815115 ,  4.259078  ,  3.5834122 ,
        0.65185696,  1.2536054 , 11.563457  , 10.591119  , 36.16629   ,
       20.65797   , 17.344439  ,  1.7194339 ,  2.6499515 ,  1.7756646 ,
        7.849437  ,  6.312393  , 12.655462  , 36.371975  , 29.804617  ]
    a = [ 9.560339 , 14.32865  ,  8.151008 ,  2.67406  ,  1.4045964,
        0.527787 ,  1.0504515,  2.610892 , 10.499335 , 15.04245  ,
       13.404196 , 20.579058 ,  2.5307505,  3.1454   ,  5.821699 ,
        4.5464473,  5.486375 ,  9.605737 , 22.239016 , 23.16811  ]
    a = [24.556381  , 78.71643   , 19.938538  ,  0.75483584,  3.4850702 ,
        1.6270295 ,  2.8741934 ,  1.6323117 ,  1.350555  ,  1.8206675 ,
        3.420386  ,  1.6071559 ,  3.6221714 ,  8.379961  , 13.977667  ,
       15.347112  ,  6.175599  ,  1.2197634 ,  3.1367052 ,  0.721017  ]
    # a = [16.149893 , 61.84469  , 14.411461 ,  1.7398348,  0.6888891,
    #     0.9446828,  2.3051815,  1.5879819,  1.2384998,  2.5209084,
    #     3.5026205,  1.0705074,  3.3557096,  7.305173 , 18.945673 ,
    #    12.923474 ,  3.2917812,  1.3232207,  1.5054771,  1.2052197]
    # a = [14.409899 ,  9.094438 ,  9.033749 ,  2.972072 ,  1.3904873,
    #     2.0952   ,  1.8999062,  6.641684 ,  1.9475664, 45.49759  ,
    #    13.806563 ,  4.564625 ,  2.6910293,  3.0519383,  5.6852283,
    #     9.018049 ,  8.699738 ,  4.63656  , 45.138477 , 42.802498 ]
    # a = [ 9.198381 , 13.399701 ,  9.830626 ,  2.5507283,  1.2544044,
    #     0.5346969,  2.2840075,  2.3242886,  1.9153802, 27.213942 ,
    #    22.73147  , 23.181873 ,  2.46287  ,  3.0273473,  5.735209 ,
    #     6.897217 ,  1.5586228,  1.0992992, 29.670061 , 30.54857  ]
    b = [115.85728  ,  13.178285 ,  47.227776 ,   3.231017 ,   1.4479513,
         1.1272107,  20.206474 ,   4.75829  ,  11.908666 ,  61.677025 ,
        10.947419 ,  40.254772 ,   2.8156958,   3.2820308,   6.028998 ,
         8.639017 ,   2.9387176,  12.271263 ,  65.85747  ,  11.751841 ]
    # a = [3.47617209e-01, 2.19977093e+00, 5.44331646e+00, 8.02162743e+00,
    #    3.82593131e+00, 1.37508404e+00, 4.10550207e-01, 1.48171425e-01,
    #    7.35632852e-02, 4.99118492e-02, 7.93359056e-02, 1.00478016e-01,
    #    1.42871663e-01, 1.58616424e-01, 1.31176278e-01, 9.92783234e-02,
    #    1.36963082e-02, 8.28585029e-03, 2.81755906e-03, 2.90522981e-03]
    # a = [0.32552913, 1.9445497 , 3.517058  , 3.867374  , 2.234789  ,
    #    1.3863355 , 0.7431453 , 0.4261789 , 0.24378935, 0.10051128,
    #    0.06644867, 0.04677694, 0.05422506, 0.04133717, 0.03463094,
    #    0.02305154, 0.01103014, 0.00782117, 0.00573242, 0.00517267]
    # a = [4.9111858e-01, 3.6719277e+00, 6.5314832e+00, 5.8773046e+00,
    #    2.7435548e+00, 1.0279704e+00, 3.3639520e-01, 1.5329474e-01,
    #    7.2476700e-02, 3.4249585e-02, 6.2258601e-02, 1.4260831e-01,
    #    1.4157829e-01, 1.0874553e-01, 8.5141815e-02, 6.5885246e-02,
    #    1.7112743e-02, 1.1381598e-02, 7.2019957e-03, 6.0972753e-03]
    a = [3.0875868e-01, 1.7475446e+00, 3.7669091e+00, 7.3978157e+00,
       3.9320047e+00, 1.5487790e+00, 3.8943410e-01, 1.2493789e-01,
       4.6345048e-02, 2.5492188e-02, 2.8380152e-02, 9.7640924e-02,
       1.2011365e-01, 1.2518139e-01, 9.6632972e-02, 6.4182714e-02,
       1.2226980e-02, 6.6843973e-03, 2.7434817e-03, 1.4130084e-03]
    # a = [2.63052553e-01, 1.59759629e+00, 3.48221707e+00, 5.12903309e+00,
    #    2.58074522e+00, 1.11498356e+00, 3.90125602e-01, 1.48772031e-01,
    #    7.32326359e-02, 4.43874002e-02, 4.19132225e-02, 8.20679814e-02,
    #    1.02619864e-01, 1.01044022e-01, 7.90384188e-02, 4.01885137e-02,
    #    8.24552868e-03, 5.24332747e-03, 4.22069291e-03, 3.46915284e-03]
    # a = [6.1824787e-01, 3.5208964e+00, 6.1176186e+00, 5.7595363e+00,
    #    2.6035471e+00, 1.1323165e+00, 2.8651801e-01, 8.4918529e-02,
    #    3.6044698e-02, 1.9677566e-02, 7.3727496e-02, 8.1269905e-02,
    #    1.0372775e-01, 9.1439627e-02, 5.6524392e-02, 3.4976285e-02,
    #    7.3396470e-03, 5.1560909e-03, 3.9179600e-03, 3.8947067e-03]
    b = [2.1024884e-01, 1.3849133e+00, 3.2203879e+00, 5.4749227e+00,
       2.6983831e+00, 1.0417283e+00, 3.2289204e-01, 1.2978017e-01,
       5.8048498e-02, 2.9843807e-02, 1.7936153e-02, 6.3620239e-02,
       5.9542548e-02, 6.1489858e-02, 4.9763292e-02, 6.0007952e-02,
       1.3045074e-02, 1.0522788e-02, 4.6142936e-03, 3.2719695e-03]
    
    a = np.array(a)
    a = softmax(a)
    a.sort()
    a = a[::-1]
    diff_a = []
    for i in range(len(a)-1):
        diff_a.append(a[i] - a[i+1])
    keep_a = 10
    print(a)
    print(np.mean(a))
    #print(diff_a)
    #print(sum(a))
    #print(sum(a[0:keep_a]))
    #print(sum(a[0:keep_a])/sum(a))
    #print(np.mean(a)/max(a))
    partial_sum = 0
    # for i in range(len(a)):
    #     partial_sum = partial_sum + a[i]
    #     print(i)
    #     print(partial_sum)
    #     print(partial_sum/sum(a))
    #     print("===")
    